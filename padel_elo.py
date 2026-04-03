import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import matplotlib.pyplot as plt

# Constants
DEFAULT_ELO = 1000
K_FACTOR = 32  # Configurable K-factor

def create_initial_excel(filename):
    """Create initial Excel file with required sheets and structure."""
    wb = openpyxl.Workbook()

    # Matches sheet
    ws_matches = wb.active
    ws_matches.title = "Matches"
    headers_matches = ["Date", "Location", "Player 1", "Player 2", "Player 3", "Player 4", "Winning Team (1 or 2)"]
    for col_num, header in enumerate(headers_matches, 1):
        cell = ws_matches.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Players sheet
    ws_players = wb.create_sheet("Players")
    headers_players = ["Player", "ELO", "Matches Played", "Wins", "Win Rate", "Most Frequent Teammate", "Most Frequent Opponent"]
    for col_num, header in enumerate(headers_players, 1):
        cell = ws_players.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Rankings sheet
    ws_rankings = wb.create_sheet("Rankings")
    headers_rankings = ["Rank", "Player", "ELO", "Matches Played", "Wins", "Win Rate"]
    for col_num, header in enumerate(headers_rankings, 1):
        cell = ws_rankings.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    wb.save(filename)
    print(f"Initial Excel file created: {filename}")

def load_players(ws_players):
    """Load players data from Players sheet."""
    players = {}
    for row in range(2, ws_players.max_row + 1):
        player_name = ws_players.cell(row=row, column=1).value
        if player_name:
            elo = ws_players.cell(row=row, column=2).value or DEFAULT_ELO
            matches = ws_players.cell(row=row, column=3).value or 0
            wins = ws_players.cell(row=row, column=4).value or 0
            players[player_name] = {
                'elo': float(elo),
                'matches': int(matches),
                'wins': int(wins),
                'teammates': {},  # dict of teammate: count
                'opponents': {}   # dict of opponent: count
            }
    return players

def save_players(ws_players, players):
    """Save players data to Players sheet."""
    # Clear existing data
    for row in range(2, ws_players.max_row + 1):
        for col in range(1, 8):
            ws_players.cell(row=row, column=col, value=None)

    # Write data
    row = 2
    for player, data in players.items():
        ws_players.cell(row=row, column=1, value=player)
        ws_players.cell(row=row, column=2, value=data['elo'])
        ws_players.cell(row=row, column=3, value=data['matches'])
        ws_players.cell(row=row, column=4, value=data['wins'])
        win_rate = data['wins'] / data['matches'] if data['matches'] > 0 else 0
        ws_players.cell(row=row, column=5, value=f"{win_rate:.2%}")
        
        # Most frequent teammate
        if data['teammates']:
            most_teammate = max(data['teammates'], key=data['teammates'].get)
            ws_players.cell(row=row, column=6, value=most_teammate)
        else:
            ws_players.cell(row=row, column=6, value="-")
        
        # Most frequent opponent
        if data['opponents']:
            most_opponent = max(data['opponents'], key=data['opponents'].get)
            ws_players.cell(row=row, column=7, value=most_opponent)
        else:
            ws_players.cell(row=row, column=7, value="-")
        
        row += 1

def calculate_team_rating(player1_elo, player2_elo):
    """Calculate team rating as average of players' ELO."""
    return (player1_elo + player2_elo) / 2

def expected_score(rating_a, rating_b):
    """Calculate expected score for player/team A vs B."""
    return 1 / (1 + 10 ** ((rating_b - rating_a) / 400))

def update_elo(current_elo, expected, actual, k=K_FACTOR):
    """Update ELO rating."""
    return current_elo + k * (actual - expected)

def process_matches(ws_matches, players):
    """Process all matches and update player ELOs."""
    for row in range(2, ws_matches.max_row + 1):
        # Read match data
        date = ws_matches.cell(row=row, column=1).value
        location = ws_matches.cell(row=row, column=2).value
        p1 = ws_matches.cell(row=row, column=3).value
        p2 = ws_matches.cell(row=row, column=4).value
        p3 = ws_matches.cell(row=row, column=5).value
        p4 = ws_matches.cell(row=row, column=6).value
        winning_team = ws_matches.cell(row=row, column=7).value

        if not all([p1, p2, p3, p4, winning_team]):
            continue  # Skip incomplete matches

        # Ensure players exist
        for p in [p1, p2, p3, p4]:
            if p not in players:
                players[p] = {'elo': DEFAULT_ELO, 'matches': 0, 'wins': 0, 'teammates': {}, 'opponents': {}}

        # Get current ELOs
        elo1 = players[p1]['elo']
        elo2 = players[p2]['elo']
        elo3 = players[p3]['elo']
        elo4 = players[p4]['elo']

        # Calculate team ratings
        team1_rating = calculate_team_rating(elo1, elo2)
        team2_rating = calculate_team_rating(elo3, elo4)

        # Expected scores
        exp_team1 = expected_score(team1_rating, team2_rating)
        exp_team2 = 1 - exp_team1

        # Actual scores
        if winning_team == 1:
            actual_team1, actual_team2 = 1, 0
        elif winning_team == 2:
            actual_team1, actual_team2 = 0, 1
        else:
            continue  # Invalid winning team

        # Update ELOs
        new_elo1 = update_elo(elo1, exp_team1, actual_team1)
        new_elo2 = update_elo(elo2, exp_team1, actual_team1)
        new_elo3 = update_elo(elo3, exp_team2, actual_team2)
        new_elo4 = update_elo(elo4, exp_team2, actual_team2)

        # Update player data
        players[p1]['elo'] = new_elo1
        players[p2]['elo'] = new_elo2
        players[p3]['elo'] = new_elo3
        players[p4]['elo'] = new_elo4

        for p in [p1, p2, p3, p4]:
            players[p]['matches'] += 1
            if (p in [p1, p2] and winning_team == 1) or (p in [p3, p4] and winning_team == 2):
                players[p]['wins'] += 1

        # Update teammates and opponents
        team1 = [p1, p2]
        team2 = [p3, p4]
        for p in team1:
            for tm in team1:
                if tm != p:
                    players[p]['teammates'][tm] = players[p]['teammates'].get(tm, 0) + 1
            for op in team2:
                players[p]['opponents'][op] = players[p]['opponents'].get(op, 0) + 1
        for p in team2:
            for tm in team2:
                if tm != p:
                    players[p]['teammates'][tm] = players[p]['teammates'].get(tm, 0) + 1
            for op in team1:
                players[p]['opponents'][op] = players[p]['opponents'].get(op, 0) + 1

def generate_rankings(ws_rankings, players):
    """Generate rankings sheet."""
    # Clear existing data
    for row in range(2, ws_rankings.max_row + 1):
        for col in range(1, 7):
            ws_rankings.cell(row=row, column=col, value=None)

    # Sort players by ELO descending
    sorted_players = sorted(players.items(), key=lambda x: x[1]['elo'], reverse=True)

    row = 2
    for rank, (player, data) in enumerate(sorted_players, 1):
        ws_rankings.cell(row=row, column=1, value=rank)
        ws_rankings.cell(row=row, column=2, value=player)
        ws_rankings.cell(row=row, column=3, value=data['elo'])
        ws_rankings.cell(row=row, column=4, value=data['matches'])
        ws_rankings.cell(row=row, column=5, value=data['wins'])
        win_rate = data['wins'] / data['matches'] if data['matches'] > 0 else 0
        ws_rankings.cell(row=row, column=6, value=f"{win_rate:.2%}")
        row += 1

def generate_elo_chart(players, filename="elo_rankings.png"):
    """Generate a bar chart of current ELO rankings."""
    if not players:
        return
    
    sorted_players = sorted(players.items(), key=lambda x: x[1]['elo'], reverse=True)
    names = [p[0] for p in sorted_players[:10]]  # Top 10
    elos = [p[1]['elo'] for p in sorted_players[:10]]
    
    plt.figure(figsize=(10, 6))
    plt.bar(names, elos, color='skyblue')
    plt.xlabel('Players')
    plt.ylabel('ELO Rating')
    plt.title('Top 10 Padel Players ELO Rankings')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig(filename)
    plt.close()
    print(f"ELO rankings chart saved as {filename}")

def main():
    filename = "padel_matches.xlsx"
    
    if not os.path.exists(filename):
        create_initial_excel(filename)
        print("Please add match data to the 'Matches' sheet and run the script again to calculate ELO rankings.")
        return

    wb = openpyxl.load_workbook(filename)
    ws_matches = wb["Matches"]
    ws_players = wb["Players"]
    ws_rankings = wb["Rankings"]

    players = load_players(ws_players)
    process_matches(ws_matches, players)
    save_players(ws_players, players)
    generate_rankings(ws_rankings, players)
    generate_elo_chart(players)

    wb.save(filename)
    print(f"ELO rankings updated in {filename}")
    print("Top 3 players:")
    sorted_players = sorted(players.items(), key=lambda x: x[1]['elo'], reverse=True)
    for i, (player, data) in enumerate(sorted_players[:3], 1):
        print(f"{i}. {player}: {data['elo']:.0f} ELO")

if __name__ == "__main__":
    main()